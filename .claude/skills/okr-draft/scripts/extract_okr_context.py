#!/usr/bin/env python3
"""OKRシート(xlsx)から対象人物のコンテキストを抽出する。

使い方:
    python extract_okr_context.py --file <xlsx> --sheet <name> --person <name>
    python extract_okr_context.py --file <xlsx> --list-sheets

stdoutにJSONを出力。Claudeはそれをパースしてドラフト生成に利用する。

シート構造の前提（OKRシート_旧名_目標＆ウェイト整理.xlsx 由来）:
- B列: 部門/上位ラベル(FY/Q/ユニットOKR等)
- C列: 担当者名(個人セクションのトップ行)
- D列: Objectives(縦結合で個人範囲をまとめる)
- E列: Key Results(個人の各KRごと1行)
- F列: 期日 / G列: ウェイト / H列: 備考 / I列: 前提条件 / J列: メモ
- Q列: 最終進捗率 / T列: 最終着地想定 / Y列: 最終GOOD / Z列: 最終MORE
- 個人のC/D/K/N/Q列は縦結合
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl が必要です: pip install openpyxl\n")
    sys.exit(1)


COLS = {
    "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7,
    "H": 8, "I": 9, "J": 10, "Q": 17, "T": 20, "Y": 25, "Z": 26,
}


def col_letter_to_idx(letter: str) -> int:
    return COLS[letter]


def find_person_row_range(ws, person_name: str) -> tuple[int, int] | None:
    """C列の結合セルから対象人物の行範囲を特定。

    完全一致を優先。なければ部分一致（姓または名）でフォールバック。
    """
    candidates = []
    for m in ws.merged_cells.ranges:
        b = m.bounds  # (min_col, min_row, max_col, max_row)
        if b[0] == COLS["C"] and b[2] == COLS["C"]:
            v = ws.cell(row=b[1], column=COLS["C"]).value
            if not v:
                continue
            v_str = str(v).strip()
            if v_str == person_name.strip():
                return (b[1], b[3])
            if person_name.strip() in v_str or v_str in person_name.strip():
                candidates.append((b[1], b[3], v_str))

    if len(candidates) == 1:
        return (candidates[0][0], candidates[0][1])
    if len(candidates) > 1:
        sys.stderr.write(
            f"対象人物名 '{person_name}' に該当する候補が複数あります:\n"
        )
        for s, e, v in candidates:
            sys.stderr.write(f"  R{s}-R{e}: {v}\n")
        return None
    return None


def extract_person_section(ws, start: int, end: int) -> dict:
    """指定行範囲のObjective/KR一覧と既存記入値を抽出。"""
    section: dict = {
        "row_range": [start, end],
        "objective": ws.cell(row=start, column=COLS["D"]).value,
        "krs": [],
    }
    for r in range(start, end + 1):
        kr = ws.cell(row=r, column=COLS["E"]).value
        if kr is None:
            continue
        f_val = ws.cell(row=r, column=COLS["F"]).value
        section["krs"].append({
            "row": r,
            "kr": str(kr),
            "due": str(f_val) if f_val else None,
            "weight": ws.cell(row=r, column=COLS["G"]).value,
            "remark": ws.cell(row=r, column=COLS["H"]).value,
            "premise": ws.cell(row=r, column=COLS["I"]).value,
            "memo": ws.cell(row=r, column=COLS["J"]).value,
            "final_progress": ws.cell(row=r, column=COLS["Q"]).value,
            "final_landing": ws.cell(row=r, column=COLS["T"]).value,
            "final_good": ws.cell(row=r, column=COLS["Y"]).value,
            "final_more": ws.cell(row=r, column=COLS["Z"]).value,
        })
    return section


def extract_parent_okrs(ws) -> list[dict]:
    """B列に `FY_`, `Q_`, または末尾が `OKR` のラベルを持つ行を親OKRとして抽出。

    親OKRの "範囲" には2種類ある:
    - row_end: B列(またはC列)の結合範囲全体。下位メンバーを含む（人物の所属判定に使用）
    - kr_row_end: 次にC列が埋まる(=次のメンバー or ラベル)行の手前まで。親OKR本体のKR範囲
    """
    parents = []
    last_row = ws.max_row
    for r in range(1, last_row + 1):
        b_val = ws.cell(row=r, column=COLS["B"]).value
        c_val = ws.cell(row=r, column=COLS["C"]).value
        if not b_val:
            continue
        b_str = str(b_val)
        is_parent = (
            b_str.startswith("FY_")
            or b_str.startswith("Q_")
            or "OKR" in b_str
            or (c_val is not None and "ユニットOKR" in str(c_val))
        )
        if not is_parent:
            continue

        # 親OKR全体の範囲（=メンバーセクションを含む結合範囲）
        person_end = r
        for m in ws.merged_cells.ranges:
            bb = m.bounds
            target_col = COLS["B"] if b_val else COLS["C"]
            if bb[0] == target_col and bb[2] == target_col and bb[1] == r:
                person_end = bb[3]
                break

        # 親OKR本体のKR範囲: r+1から見て、C列が埋まる最初の行の前まで
        kr_end = person_end
        for rr in range(r + 1, person_end + 1):
            if ws.cell(row=rr, column=COLS["C"]).value is not None:
                kr_end = rr - 1
                break

        parents.append({
            "label": b_str,
            "sub_label": str(c_val) if c_val else None,
            "row": r,
            "row_end": person_end,
            "kr_row_end": kr_end,
            "objective": ws.cell(row=r, column=COLS["D"]).value,
            "krs": _collect_krs(ws, r, kr_end),
        })
    return parents


def _collect_krs(ws, start: int, end: int) -> list[dict]:
    out = []
    for r in range(start, end + 1):
        kr = ws.cell(row=r, column=COLS["E"]).value
        if kr is None:
            continue
        out.append({
            "row": r,
            "kr": str(kr),
            "due": str(ws.cell(row=r, column=COLS["F"]).value) if ws.cell(row=r, column=COLS["F"]).value else None,
            "weight": ws.cell(row=r, column=COLS["G"]).value,
        })
    return out


def find_unit_for_person(parent_okrs: list[dict], person_row: int) -> dict | None:
    """対象人物が所属するユニットOKRを特定。
    人物の行が parents の row_end 範囲内にある最後のユニットOKRを返す。
    """
    candidate = None
    for p in parent_okrs:
        if p["row"] <= person_row <= p["row_end"]:
            if "ユニット" in (p.get("sub_label") or ""):
                candidate = p
    return candidate


def extract_peer_samples(ws, parent_unit: dict | None, target_row_range: tuple[int, int], max_peers: int = 2) -> list[dict]:
    """対象人物と同じユニット内の他メンバーの記入例を抽出。"""
    if parent_unit is None:
        return []
    peers = []
    seen_ranges = set()
    for m in ws.merged_cells.ranges:
        b = m.bounds
        if b[0] != COLS["C"] or b[2] != COLS["C"]:
            continue
        if not (parent_unit["row"] <= b[1] <= parent_unit["row_end"]):
            continue
        if (b[1], b[3]) == tuple(target_row_range):
            continue
        if (b[1], b[3]) in seen_ranges:
            continue
        seen_ranges.add((b[1], b[3]))

        name = ws.cell(row=b[1], column=COLS["C"]).value
        if not name:
            continue
        # KRが1個も入っていないメンバーは参考にならないのでスキップ
        section = extract_person_section(ws, b[1], b[3])
        if not section["krs"]:
            continue
        peers.append({
            "name": str(name).strip(),
            **section,
        })
        if len(peers) >= max_peers:
            break
    return peers


_QUARTER_RE = re.compile(r"^(\d{4})_Q([1-4])$")


def list_past_quarter_sheets(all_sheets: list[str], current: str, max_count: int = 2) -> list[str]:
    """current より時系列で過去のシートを max_count 個返す（新しい順）。"""
    m = _QUARTER_RE.match(current)
    if not m:
        return []
    cur_year, cur_q = int(m.group(1)), int(m.group(2))
    cur_idx = cur_year * 4 + cur_q

    pasts = []
    for s in all_sheets:
        m2 = _QUARTER_RE.match(s)
        if not m2:
            continue
        y, q = int(m2.group(1)), int(m2.group(2))
        idx = y * 4 + q
        if idx < cur_idx:
            pasts.append((idx, s))
    pasts.sort(reverse=True)
    return [s for _, s in pasts[:max_count]]


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    p.add_argument("--file", required=True, help="xlsxファイルパス")
    p.add_argument("--sheet", help="対象シート名 (例: 2026_Q2)")
    p.add_argument("--person", help="対象人物名 (例: 唐子 征久)")
    p.add_argument("--list-sheets", action="store_true", help="シート一覧のみ出力")
    p.add_argument("--past-count", type=int, default=2, help="過去Quarterを何個取得するか")
    args = p.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        sys.stderr.write(f"ファイルが存在しません: {xlsx_path}\n")
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if args.list_sheets:
        json.dump({"sheets": wb.sheetnames}, sys.stdout, ensure_ascii=False, indent=2)
        sys.stdout.write("\n")
        return 0

    if not args.sheet:
        sys.stderr.write("--sheet が必要です（または --list-sheets でシート一覧取得）\n")
        return 1
    if args.sheet not in wb.sheetnames:
        sys.stderr.write(f"シート '{args.sheet}' が存在しません\n")
        sys.stderr.write(f"利用可能: {wb.sheetnames}\n")
        return 1
    if not args.person:
        sys.stderr.write("--person が必要です\n")
        return 1

    ws = wb[args.sheet]
    person_range = find_person_row_range(ws, args.person)
    if person_range is None:
        sys.stderr.write(f"対象人物 '{args.person}' が見つかりません（C列の結合セル）\n")
        return 2

    target_section = extract_person_section(ws, *person_range)
    parent_okrs = extract_parent_okrs(ws)
    parent_unit = find_unit_for_person(parent_okrs, person_range[0])
    peer_samples = extract_peer_samples(ws, parent_unit, person_range)

    past_quarters = []
    past_sheet_names = list_past_quarter_sheets(wb.sheetnames, args.sheet, args.past_count)
    for s_name in past_sheet_names:
        ws_past = wb[s_name]
        past_range = find_person_row_range(ws_past, args.person)
        if past_range is None:
            continue
        past_section = extract_person_section(ws_past, *past_range)
        past_quarters.append({
            "sheet": s_name,
            **past_section,
        })

    result = {
        "file": str(xlsx_path),
        "sheet": args.sheet,
        "person": args.person,
        "target_row_range": list(person_range),
        "target_existing": target_section,
        "past_quarters": past_quarters,
        "parent_okrs": parent_okrs,
        "parent_unit_for_person": parent_unit,
        "peer_samples": peer_samples,
    }
    json.dump(result, sys.stdout, ensure_ascii=False, indent=2, default=str)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
